# -*- coding: utf-8 -*-

import networkx as nx
import openpyxl as px
from nodes import genes_classify

book = px.load_workbook(r'C:\users\administrator\desktop\all_data.xlsx')
data = book[book.sheetnames[0]]


def index(gene, sample):
    column = list()
    row = list()
    for c in data['A']:
        if c.value:
            column.append(c.value)  # genes
    for r in data[1]:
        if r.value:
            row.append(r.value)  # sample

    row_index = dict()
    column_index = dict()
    for ci in enumerate(column):
        column_index[ci[1]] = ci[0] + 2
    for ri in enumerate(row):
        row_index[ri[1]] = ri[0] + 2
    return data.cell(column_index[gene], row_index[sample]).value


def node():
    gene_node = [x.value for x in data['A'] if x.value]
    sample_node = [x.value for x in data[1] if x.value]
    return sample_node, gene_node[:228]


def draw():
    g = nx.Graph()
    for n1 in node()[0]:
        g.add_node(n1, flag='sample')
    for n2 in node()[1]:
        g.add_node(n2, flag=genes_classify()[n2])

    for e1 in node()[0]:  # sample
        for e2 in node()[1]:  # gene
            g.add_edge(e1, e2, weight=100 * index(e2, e1))
    return g


if __name__ == '__main__':
    graph = draw()
    nx.write_gml(graph, r"C:\users\administrator\\desktop\network\genes_percent_2.gml")
    # print(index('aacC1', 'CMU'))
