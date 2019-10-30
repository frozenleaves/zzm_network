# -*- coding: utf-8 -*-


import openpyxl

r_data = openpyxl.load_workbook('C:\\users\\administrator\\desktop\\rb.xlsx')
p_data = openpyxl.load_workbook('C:\\users\\administrator\\desktop\\pb.xlsx')


def index(name):
    key_cell = r_data[r_data.sheetnames[0]]['A']
    keys = list()
    for key in key_cell:
        keys.append(key.value)
    keys.pop(0)

    index_dict = dict()

    for i in enumerate(keys):
        index_dict[i[1]] = i[0] + 2
    return index_dict[name]


def select(row, column, flag: 'if flag = 0, return r data, else, return p data' = 0):
    if flag:
        return p_data[p_data.sheetnames[0]].cell(index(row), index(column)).value
    else:
        return r_data[r_data.sheetnames[0]].cell(index(row), index(column)).value


def nodes():
    key_cell = r_data[r_data.sheetnames[0]]['A']
    keys = list()
    for key in key_cell:
        keys.append(key.value)
    keys.pop(0)
    return keys


if __name__ == '__main__':
    select(index("aac(6')-Ib(aka aacA4)-01"), index("aac(6')-Ib(aka aacA4)-02"))
