# -*- coding: utf-8 -*-

import openpyxl

file1 = r"C:\users\administrator\desktop\all_data.xlsx"
file2 = r"C:\users\administrator\desktop\分类.xlsx"
book1 = openpyxl.load_workbook(file1)  # 汇总数据
book2 = openpyxl.load_workbook(file2)  # 分类数据

sheets = book1.sheetnames[0]
x = book1[sheets]


def get_row_data(sheet, column) -> 'return: each column data':
    column_data = list()
    for j in range(len(sheet['1'])):
        column_data.append(sheet[column][j].value)
    column_data.pop(0)
    return column_data


def get_column_data(sheet, row) -> 'return: each row data':
    row_data = list()
    for index in range(len(sheet['A'])):
        if sheet[row][index].value:
            row_data.append(sheet[row][index].value)

    return row_data


def get_cell_data(genes, sample):
    """

    :param sample: 10 sample names
    :param genes: 228 genes names
    :return: cell value for params
    """

    sample_dict = dict()
    genes_dict = dict()

    for i in enumerate(get_row_data(x, 1)):
        sample_dict[i[1]] = i[0] + 2

    for j in enumerate(get_column_data(x, 'A')):
        genes_dict[j[1]] = j[0] + 2

    data = book1[book1.sheetnames[0]].cell(genes_dict[genes], sample_dict[sample]).value
    return data


if __name__ == '__main__':
    # get_cell_data('16S rRNA', 'CMU')
    pass
