# -*- coding: utf-8 -*-

from copy import deepcopy
import openpyxl as pl
import read_excel as re

all_data_book = pl.load_workbook(r"C:\users\administrator\desktop\all_data.xlsx")
classify_book = pl.load_workbook(r"C:\users\administrator\desktop\分类.xlsx")
phylum_book = pl.load_workbook(r"C:\users\administrator\desktop\phylum.xlsx")
genes_keys_sheet = all_data_book[all_data_book.sheetnames[0]]
phylum_key_sheet = phylum_book[phylum_book.sheetnames[0]]


def load_key_data():
    """
    加载所需数据
    :return: 数据集列表
    """

    genes_key = re.get_column_data(genes_keys_sheet, 'A')
    sample_key = re.get_row_data(genes_keys_sheet, 1)
    classify_key = classify_book.sheetnames

    phylum_key = list()
    for i in phylum_key_sheet[1]:
        phylum_key.append(i.value)
    phylum_key.pop(0)

    return genes_key, sample_key, classify_key


def load_copy_number():
    """
    加载每个样本的每个基因拷贝数
    :return: 数据集列表，包含样本，抗性基因，抗性基因拷贝数三元元组
    """
    key = load_key_data()
    msg_list = list()

    for sample in key[1]:
        for genes in key[0]:
            msg_list.append((sample, genes, re.get_cell_data(genes, sample)))
    return msg_list


def sample_msg():
    """

    :return: dict:{
    'sample_name': [(gene_name, copy_number)]
    }
    """
    msg_dict = dict()
    for key in load_key_data()[1]:
        genes_list = []
        for msg in load_copy_number():
            if msg[0] == key:
                if msg[2]:
                    genes_list.append((msg[1], msg[2]))
        msg_dict[key] = deepcopy(genes_list)
        genes_list.clear()
    return msg_dict


def phylum_msg():
    """
    :return:
    {
        'sample_name':[(phylum, abundance)]
        }
    """
    phylum_dict = dict()
    for sample in load_key_data()[1]:
        dict_value = list()
        for phylum_data in set_phylum_data():
            if sample == phylum_data[0]:
                if phylum_data[2]:
                    dict_value.append((phylum_data[1], phylum_data[2]))
        phylum_dict[sample] = deepcopy(dict_value)
        dict_value.clear()

    return phylum_dict


def genes_classify():
    """
    抗性基因所属分类
    :return: 类名-包含抗性基因集合的字典
    """
    classify = classify_book.sheetnames
    class_dict = dict()
    for i in classify:
        for j in re.get_column_data(classify_book[i], 'A'):
            class_dict[j] = i
    return class_dict


def phylum_cell_data(sample_, phylum_):
    """
    门一级物种丰度数据
    :return: 样本每个门的物种丰度元组
    """
    sample_index = dict()
    phylum_index = dict()
    for sample in enumerate(re.get_column_data(phylum_key_sheet, 'A')):
        sample_index[sample[1]] = sample[0] + 2
    for phylum in enumerate(re.get_row_data(phylum_key_sheet, '1')):
        phylum_index[phylum[1]] = phylum[0] + 2

    return phylum_key_sheet.cell(sample_index[sample_], phylum_index[phylum_]).value


def set_phylum_data():
    phylum_data_list = list()
    for phylum in re.get_row_data(phylum_key_sheet, '1'):
        phylum_data_list.append(phylum)
    return phylum_data_list


def set_genes_nodes(name, sort):
    """
    抗性基因节点
    包含属性：
        抗性基因名称
        抗性基因相对拷贝数
        抗性基因所属分类
    :return: node object
    """

    class GenesNode(object):
        def __init__(self):
            self.name = name
            self.sort = sort

        def __str__(self):
            return self.name

        def __repr__(self):
            return self.name

    return GenesNode()


def set_phylum_nodes(name):
    """
    门一级物种节点：
    包含属性：
        物种名称
        物种丰度
    :return: node object
    """

    class PhylumNode(object):
        def __init__(self):
            # self.sample = sample
            self.name = name

        def __str__(self):
            return self.name

        def __repr__(self):
            return self.name

    return PhylumNode()


def set_sample_nodes(name):
    """
    样本节点
    包含属性：
        样本名称

    :return: nodes object
    """

    class SampleNode(object):
        def __init__(self):
            self.name = name

        def __str__(self):
            return self.name

        def __repr__(self):
            return self.name

    return SampleNode()


def create_genes_node():
    """
    创建节点对象
    :return:  genes node instance list
    """
    genes_sort = genes_classify()
    genes_nodes = list()

    for genes in load_key_data()[0]:
        genes_nodes.append(set_genes_nodes(genes, genes_sort[genes]))
    # print(genes_sort["aac(6')0Ib(aka aacA4)001"])

    return genes_nodes


def create_sample_node():
    """

    :return: sample node instances
    """
    sample_nodes = list()
    for sample in load_key_data()[1]:
        # name = item[0]
        # gene_msgs = item[1]
        # phylum_msgs = phylum_msg()[name]
        sample_nodes.append(set_sample_nodes(sample))
    return sample_nodes


def create_phylum_node():
    """

    :return:
    """
    phylum_nodes = list()
    for phylum in set_phylum_data():
        phylum_nodes.append(set_phylum_nodes(phylum))
    return phylum_nodes


def main():
    pass


if __name__ == '__main__':
    main()
